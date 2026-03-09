from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from datetime import datetime
import uuid
import logging
from pathlib import Path
from typing import List, Dict
import json

import httpx
from pydantic import BaseModel, Field

from config import (
    FRONTEND_URLS,
    API_TITLE,
    API_VERSION,
    API_DESCRIPTION,
    UPLOAD_FOLDER,
    MAX_FILE_SIZE,
    TEMP_FOLDER,
    BASE_DIR,
    GEMINI_API_KEY,
    GEMINI_MODEL,
)
from firebase_service import OrcamentoFirestore
from budget_parser import BudgetParser

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

_GEMINI_CANDIDATE_MODELS = [
    GEMINI_MODEL,
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
    "gemini-1.5-flash-latest",
    "gemini-1.5-pro-latest",
    "gemini-pro",
]


async def _call_gemini_generate_content(request_body: dict, timeout_seconds: float = 45.0):
    attempted_models = []
    last_error_body = ""

    models_to_try = []
    for model in _GEMINI_CANDIDATE_MODELS:
        if model and model not in models_to_try:
            models_to_try.append(model)

    async with httpx.AsyncClient(timeout=timeout_seconds) as client:
        for model in models_to_try:
            attempted_models.append(model)
            url = (
                f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
                f"?key={GEMINI_API_KEY}"
            )
            response = await client.post(url, json=request_body)

            if response.status_code < 400:
                logger.info(f"✅ Gemini respondeu com modelo: {model}")
                return response.json(), model

            last_error_body = response.text
            logger.warning(
                f"⚠️ Gemini falhou com modelo {model} (status {response.status_code})."
            )

            if response.status_code == 404:
                continue

            if response.status_code == 429:
                raise HTTPException(
                    status_code=429,
                    detail="Quota da API do Gemini excedida. Aguarde alguns minutos ou troque a chave/modelo.",
                )

            raise HTTPException(
                status_code=502,
                detail=f"Erro na API do Gemini: {response.text}",
            )

    raise HTTPException(
        status_code=502,
        detail=(
            "Nenhum modelo Gemini compatível respondeu ao generateContent. "
            f"Modelos tentados: {', '.join(attempted_models)}. "
            f"Último erro: {last_error_body}"
        ),
    )

# Cache em memória para modo offline (dados temporários)
_OFFLINE_CACHE = {}

# FastAPI app
app = FastAPI(
    title=API_TITLE,
    version=API_VERSION,
    description=API_DESCRIPTION,
)

# CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_URLS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============== STATIC FILES (FRONTEND) ==============
# Servir arquivos estáticos do frontend build
FRONTEND_DIST = BASE_DIR.parent / "frontend" / "dist"
if FRONTEND_DIST.exists():
    logger.info(f"✅ Frontend dist encontrado: {FRONTEND_DIST}")
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIST / "assets", check_dir=False), name="assets")
else:
    logger.warning(f"⚠️  Frontend dist não encontrado: {FRONTEND_DIST}")

# ============== HEALTH CHECK ==============
@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "online",
        "timestamp": datetime.now().isoformat(),
        "version": API_VERSION,
    }

# ============== TEST ==============
@app.get("/api/test")
async def test_endpoint():
    """Test endpoint to verify API is working"""
    return {
        "message": "✅ Backend está funcionando!",
        "timestamp": datetime.now().isoformat(),
    }

# ============== AI STANDARDIZATION ==============
class AIItem(BaseModel):
    descricao: str
    quantidade: float
    unidade: str
    valor_unitario: float
    valor_total: float


class AIStandardizeRequest(BaseModel):
    items: List[AIItem] = Field(default_factory=list)


@app.post("/api/ai/standardize")
async def ai_standardize_items(payload: AIStandardizeRequest):
    if not GEMINI_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Recurso de IA indisponível: chave do Gemini não configurada",
        )

    system_message = (
        "Você é um assistente de normalização de dados de orçamento. "
        "Padronize descrições e unidades de medida, mantendo quantidades e valores. "
        "Retorne apenas JSON válido com uma lista de itens."
    )
    user_message = {
        "tarefa": "padronizar_itens",
        "regras": {
            "unidades": ["un", "m", "m2", "m3", "kg", "t", "l"],
            "manter_campos": ["quantidade", "valor_unitario", "valor_total"],
        },
        "items": [item.model_dump() for item in payload.items],
        "formato_retorno": {
            "items": [
                {
                    "descricao": "string",
                    "quantidade": 0,
                    "unidade": "string",
                    "valor_unitario": 0,
                    "valor_total": 0,
                }
            ]
        },
    }

    try:
        request_body = {
            "contents": [
                {
                    "parts": [
                        {"text": system_message},
                        {"text": json.dumps(user_message)}
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.1,
                "topK": 40,
                "topP": 0.95,
            }
        }

        response_data, _model_used = await _call_gemini_generate_content(
            request_body,
            timeout_seconds=30.0,
        )
        if "candidates" not in response_data or not response_data["candidates"]:
            raise ValueError("Resposta vazia do Gemini")
        
        content = response_data["candidates"][0]["content"]["parts"][0]["text"].strip()
        
        # Extrair JSON da resposta
        if content.startswith("```"):
            content = content.strip("`")
            if content.startswith("json"):
                content = content[4:].strip()

        parsed = json.loads(content)
        items = parsed.get("items") if isinstance(parsed, dict) else parsed
        if not isinstance(items, list):
            raise ValueError("Resposta de IA inválida")

        return {
            "status": "success",
            "items": items,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"❌ Erro ao padronizar itens com IA: {exc}")
        raise HTTPException(
            status_code=500,
            detail="Erro ao processar padronização com IA",
        )

# ============== UPLOAD PDF ==============
@app.post("/api/upload")
async def upload_pdf(file: UploadFile = File(...)):
    """
    Upload de arquivo PDF
    
    Returns:
        {
            "status": "success",
            "upload_id": "uuid",
            "filename": "documento.pdf",
            "size": 1234567,
            "message": "Arquivo recebido com sucesso"
        }
    """
    try:
        # Validar tipo de arquivo
        if file.content_type != "application/pdf":
            raise HTTPException(
                status_code=400,
                detail="❌ Apenas arquivos PDF são permitidos",
            )
        
        # Validar tamanho (50MB)
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"❌ Arquivo muito grande. Máximo: 50MB. Seu arquivo tem: {len(contents) / 1024 / 1024:.2f}MB",
            )
        
        # Gerar ID único
        upload_id = str(uuid.uuid4())
        file_path = UPLOAD_FOLDER / f"{upload_id}_{file.filename}"
        
        # Salvar arquivo
        with open(file_path, "wb") as buffer:
            buffer.write(contents)
        
        logger.info(f"✅ PDF salvo: {file_path} ({len(contents) / 1024 / 1024:.2f}MB)")
        
        return {
            "status": "success",
            "upload_id": upload_id,
            "filename": file.filename,
            "size": len(contents),
            "message": "✅ Arquivo recebido com sucesso",
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro no upload: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao fazer upload: {str(e)}",
        )

# ============== EXTRACT PDF ==============
@app.post("/api/extract")
async def extract_pdf(upload_id: str):
    """
    Extrai tabelas do PDF usando pdfplumber
    Salva dados no Firestore e deleta arquivo PDF
    
    Args:
        upload_id: ID retornado pelo endpoint /api/upload
    
    Returns:
        {
            "status": "success",
            "upload_id": "uuid",
            "document_id": "firestore_doc_id",
            "tables_found": 1,
            "tables": [...]
        }
    """
    try:
        if not pdfplumber:
            raise HTTPException(
                status_code=500,
                detail="pdfplumber não está instalado",
            )
        
        # Encontrar arquivo
        files = list(UPLOAD_FOLDER.glob(f"{upload_id}_*"))
        if not files:
            raise HTTPException(
                status_code=404,
                detail=f"❌ Upload não encontrado: {upload_id}",
            )
        
        file_path = files[0]
        filename = file_path.name
        
        # Extrair tabelas
        tables = []
        try:
            with pdfplumber.open(file_path) as pdf:
                logger.info(f"📄 Processando PDF: {len(pdf.pages)} página(s)")
                
                for page_num, page in enumerate(pdf.pages):
                    logger.info(f"  Página {page_num + 1}: {page.width}x{page.height}")
                    
                    # Estratégia 1: extract_tables() padrão
                    page_tables = page.extract_tables()
                    
                    # Se não encontrou tabelas, tentar com configurações customizadas
                    if not page_tables:
                        logger.info(f"  Tentando extração com settings customizados...")
                        try:
                            page_tables = page.extract_tables({
                                "vertical_strategy": "text",
                                "horizontal_strategy": "text",
                                "snap_tolerance": 5,
                                "join_tolerance": 5,
                                "edge_min_length": 3,
                            })
                        except Exception as e:
                            logger.warning(f"  Erro na extração customizada: {str(e)}")
                    
                    # Se ainda não encontrou tabelas, tentar extrair texto estruturado
                    if not page_tables:
                        logger.info(f"  Tentando extração de texto estruturado...")
                        text = page.extract_text()
                        if text:
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            if lines:
                                # Criar uma "tabela" com as linhas de texto
                                page_tables = [[[line] for line in lines]]
                                logger.info(f"  Extraído {len(lines)} linhas de texto")
                    
                    if page_tables:
                        for table_idx, table in enumerate(page_tables):
                            # Processar células mescladas (None) e limpar dados
                            processed_rows = []
                            for row in table:
                                processed_row = []
                                for cell in row:
                                    # Converter None para string vazia
                                    if cell is None:
                                        processed_row.append("")
                                    # Limpar espaços extras e quebras de linha
                                    elif isinstance(cell, str):
                                        cleaned = cell.strip().replace('\n', ' ')
                                        processed_row.append(cleaned)
                                    else:
                                        processed_row.append(str(cell))
                                processed_rows.append(processed_row)
                            
                            tables.append({
                                "page": page_num + 1,
                                "table_id": f"page_{page_num}_table_{table_idx}",
                                "rows": processed_rows,
                                "original_rows": len(table),
                                "columns": len(table[0]) if table else 0
                            })
                            logger.info(f"  ✓ Tabela {table_idx + 1}: {len(processed_rows)} linhas x {len(table[0]) if table else 0} colunas")
                    else:
                        logger.warning(f"  ⚠️  Nenhuma tabela encontrada na página {page_num + 1}")
                        
        except Exception as e:
            logger.error(f"❌ Erro ao extrair tabelas: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Erro ao extrair tabelas: {str(e)}",
            )
        
        logger.info(f"✅ {len(tables)} tabela(s) extraída(s) de {file_path}")
        
        # Parsear itens das tabelas usando parser inteligente
        parser = BudgetParser()
        parsed_data = parser.parse_all_tables(tables)
        items = parsed_data.get('items', [])
        resumo = parsed_data.get('resumo', {})
        
        logger.info(f"📊 Parser extraiu {len(items)} itens (confiança: {resumo.get('confianca', 0):.2f})")
        
        # Salvar no Firestore
        try:
            doc_id = OrcamentoFirestore.save_orcamento(
                upload_id=upload_id,
                filename=filename,
                tables=tables,
                items_data={'items': items, 'resumo': resumo}
            )
            logger.info(f"✅ Dados salvos no Firestore: {doc_id}")
        except Exception as e:
            logger.error(f"❌ Erro ao salvar no Firestore: {str(e)}")
            # Continuar mesmo se Firestore falhar, dados extraídos ainda serão retornados
            doc_id = None
        
        # Salvar em cache para modo offline
        _OFFLINE_CACHE[upload_id] = {
            "uploadId": upload_id,
            "filename": filename,
            "tables": tables,
            "items": items,
            "resumo": resumo,
            "uploadedAt": datetime.now().isoformat(),
            "extractedAt": datetime.now().isoformat(),
            "tablesFound": len(tables),
            "itemsFound": len(items),
            "status": "completed"
        }
        logger.info(f"✅ Dados salvos em cache offline: {upload_id}")
        
        # Deletar arquivo PDF
        try:
            file_path.unlink()
            logger.info(f"🗑️  PDF deletado: {file_path}")
        except Exception as e:
            logger.warning(f"⚠️  Erro ao deletar PDF: {str(e)}")
        
        return {
            "status": "success",
            "upload_id": upload_id,
            "document_id": doc_id,
            "filename": filename,
            "tables_found": len(tables),
            "items_found": len(items),
            "tables": tables,
            "items": items,
            "resumo": resumo,
            "message": "✅ Dados extraídos e processados com sucesso"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro na extração: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=str(e),
        )

# ============== ANALYZE WITH AI ==============
class AnalyzeWithAIRequest(BaseModel):
    upload_id: str
    focus: str = "budget"  # budget, items, structure, all

@app.post("/api/analyze-with-ai")
async def analyze_with_ai(payload: AnalyzeWithAIRequest):
    """
    Análise inteligente de dados extraídos com IA Gemini
    
    Usa IA para:
    - Identificar estrutura de planilha orçamentária
    - Reconhecer colunas (descrição, quantidade, unidade, valor)
    - Filtrar linhas irrelevantes (subtotais, linhas em branco)
    - Validar dados e sugerir correções
    
    Args:
        upload_id: ID retornado pelo /api/upload
        focus: Tipo de análise (budget, items, structure, all)
    
    Returns:
        {
            "status": "success",
            "upload_id": "uuid",
            "analysis": {
                "structure": {...},
                "items": [...],
                "metadata": {...}
            }
        }
    """
    try:
        if not GEMINI_API_KEY:
            raise HTTPException(
                status_code=503,
                detail="Recurso de IA indisponível: chave do Gemini não configurada",
            )
        
        # Buscar dados extraídos
        upload_data = _OFFLINE_CACHE.get(payload.upload_id)
        if not upload_data:
            raise HTTPException(
                status_code=404,
                detail=f"❌ Dados extraídos não encontrados: {payload.upload_id}",
            )
        
        tables = upload_data.get("tables", [])
        if not tables:
            raise HTTPException(
                status_code=400,
                detail="Nenhuma tabela encontrada para análise",
            )
        
        # Preparar texto das tabelas para análise
        tables_text = ""
        for table in tables:
            rows = table.get("rows", [])
            table_text = "Página {}, Tabela: {} linhas x {} colunas\n".format(
                table.get("page", "?"),
                len(rows),
                table.get("columns", "?")
            )
            for row in rows[:20]:  # Limitar a 20 linhas por tabela para análise
                table_text += " | ".join(str(cell)[:40] for cell in row) + "\n"
            tables_text += table_text + "\n---\n"
        
        # Payload para Gemini
        system_message = """Você é um especialista em análise de orçamentos e planilhas de construção civil.
Analise os dados extraídos de um PDF de orçamento e:
1. Identifique a estrutura (quais colunas representam descrição, quantidade, unidade, valor)
2. Extraia e valide os items orçamentários
3. Filtre e descarte linhas irrelevantes (subtotais, totalizações, cabeçalhos duplicados, linhas em branco)
4. Retorne apenas JSON estruturado com os dados validados

Estrutura esperada de cada item:
- descricao: string com descrição do serviço/material
- quantidade: número (pode ter decimais)
- unidade: string (un, m, m2, m3, kg, t, l, h, dia, etc)
- valor_unitario: número em reais
- valor_total: quantidade * valor_unitario

Regras:
- Descartar linhas onde descricao contém: "total", "subtotal", "suma", "resumen"
- Descartar linhas que parecem ser títulos ou seções
- Converter valores de string para número (remover R$, converter vírgula em ponto)
- Manter apenas items que tenham pelo menos: descrição e um valor numérico"""
        
        user_message = {
            "tarefa": "analisar_orcamento",
            "dados_extraidos": tables_text,
            "requisitos": {
                "validar_estrutura": True,
                "filtrar_linhas_invalidas": True,
                "identificar_colunas": True,
                "extrair_items": True
            },
            "formato_retorno": {
                "structure": {
                    "coluna_descricao": 0,
                    "coluna_quantidade": 1,
                    "coluna_unidade": 2,
                    "coluna_valor_unitario": 3,
                    "confianca": 0.95
                },
                "items": [
                    {
                        "id": "item_1",
                        "descricao": "descrição do item",
                        "quantidade": 10.0,
                        "unidade": "un",
                        "valor_unitario": 100.50,
                        "valor_total": 1005.0,
                        "validado": True,
                        "notas": ""
                    }
                ],
                "resumo": {
                    "total_items": 0,
                    "valor_total": 0.0,
                    "confianca_analise": 0.95,
                    "avisos": []
                }
            }
        }
        
        request_body = {
            "contents": [
                {
                    "parts": [
                        {"text": system_message},
                        {"text": json.dumps(user_message, ensure_ascii=False)}
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.2,  # Baixa temperatura para respostas mais determinísticas
                "topK": 40,
                "topP": 0.95,
                "maxOutputTokens": 4000,
            }
        }

        logger.info("📤 Enviando para Gemini para análise...")
        response_data, _model_used = await _call_gemini_generate_content(
            request_body,
            timeout_seconds=45.0,
        )
        if "candidates" not in response_data or not response_data["candidates"]:
            raise ValueError("Resposta vazia do Gemini")
        
        content = response_data["candidates"][0]["content"]["parts"][0]["text"].strip()
        
        # Extrair JSON da resposta
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()
        
        logger.info("✅ Resposta recebida do Gemini")
        
        # Parse JSON
        try:
            analysis = json.loads(content)
        except json.JSONDecodeError as e:
            logger.error(f"Erro ao parsear JSON: {content[:200]}")
            raise HTTPException(
                status_code=502,
                detail="IA retornou resposta inválida",
            )
        
        # Validar estrutura
        items = analysis.get("items", [])
        summary = analysis.get("resumo", {})
        structure = analysis.get("structure", {})
        
        # Enriquecer dados
        if not summary.get("valor_total"):
            summary["valor_total"] = sum(item.get("valor_total", 0) for item in items)
        if not summary.get("total_items"):
            summary["total_items"] = len(items)
        
        # Salvar análise em cache
        _OFFLINE_CACHE[payload.upload_id]["ai_analysis"] = {
            "analyzed_at": datetime.now().isoformat(),
            "structure": structure,
            "items": items,
            "summary": summary
        }
        
        logger.info(f"✅ Análise concluída: {len(items)} items reconhecidos")
        
        return {
            "status": "success",
            "upload_id": payload.upload_id,
            "analysis": {
                "structure": structure,
                "items": items,
                "summary": summary,
                "confianca_geral": summary.get("confianca_analise", 0.8)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro na análise com IA: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao analisar com IA: {str(e)}",
        )

# ============== FIRESTORE OPERATIONS ==============

@app.get("/api/orcamentos")
async def list_orcamentos():
    """
    Listar todos os orçamentos salvos no Firestore
    
    Returns:
        {
            "status": "success",
            "count": 5,
            "orcamentos": [...]
        }
    """
    try:
        orcamentos = OrcamentoFirestore.list_all_orcamentos()
        return {
            "status": "success",
            "count": len(orcamentos),
            "orcamentos": orcamentos,
        }
    except Exception as e:
        logger.error(f"❌ Erro ao listar orçamentos: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao listar orçamentos: {str(e)}",
        )

@app.get("/api/orcamentos/{upload_id}")
async def get_orcamento(upload_id: str):
    """
    Recuperar orçamento específico do Firestore
    
    Args:
        upload_id: Upload ID
    
    Returns:
        {
            "status": "success",
            "orcamento": {...}
        }
    """
    try:
        orcamento = OrcamentoFirestore.get_orcamento_by_upload_id(upload_id)
        
        if not orcamento:
            raise HTTPException(
                status_code=404,
                detail=f"❌ Orçamento não encontrado: {upload_id}",
            )
        
        return {
            "status": "success",
            "orcamento": orcamento,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao recuperar orçamento: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao recuperar orçamento: {str(e)}",
        )

# ============== CURVA ABC ==============
@app.get("/api/curva-abc/{upload_id}")
async def get_curva_abc(upload_id: str):
    """
    Calcula Curva ABC (análise de Pareto) para itens do orçamento
    
    Args:
        upload_id: Upload ID do orçamento
    
    Returns:
        {
            "status": "success",
            "items": [...],
            "summary": {...}
        }
    """
    try:
        # Buscar orçamento
        orcamento = OrcamentoFirestore.get_orcamento_by_upload_id(upload_id)
        
        # Se não encontrou no Firestore, tentar buscar do cache offline
        if not orcamento:
            orcamento = _OFFLINE_CACHE.get(upload_id)
        
        if not orcamento:
            raise HTTPException(
                status_code=404,
                detail=f"❌ Orçamento não encontrado: {upload_id}",
            )
        
        # Usar items já extraídos pelo parser (se disponíveis)
        items = orcamento.get("items", [])
        
        # Se não tem items extraídos, tentar das tabelas (fallback legado)
        if not items:
            tables = orcamento.get("tables", [])
            items = []
            
            for table in tables:
                rows = table.get("rows", [])
                
                # Pular primeira linha (cabeçalho)
                for row in rows[1:]:
                    if len(row) < 4:
                        continue
                    
                    try:
                        # Tentar extrair: descrição, quantidade, unidade, valor unitário
                        descricao = str(row[0] or "").strip()
                        quantidade_str = str(row[1] or "").strip()
                        unidade = str(row[2] or "").strip()
                        valor_str = str(row[3] or "").strip()
                        
                        if not descricao or descricao.lower() in ["total", "subtotal", ""]:
                            continue
                        
                        # Limpar e converter valores numéricos
                        quantidade = float(quantidade_str.replace(",", "."))
                        valor_unitario = float(valor_str.replace("R$", "").replace(",", ".").strip())
                        valor_total = quantidade * valor_unitario
                        
                        items.append({
                            "id": f"item_{len(items) + 1}",
                            "descricao": descricao,
                            "quantidade": quantidade,
                            "unidade": unidade,
                            "valor_unitario": valor_unitario,
                            "valor_total": valor_total,
                            "status": "validado"
                        })
                    except (ValueError, IndexError, TypeError):
                        continue
        
        if not items:
            return {
                "status": "success",
                "items": [],
                "summary": {
                    "total": 0,
                    "countA": 0,
                    "countB": 0,
                    "countC": 0,
                    "valueA": 0,
                    "valueB": 0,
                    "valueC": 0,
                    "percentA": 0,
                    "percentB": 0,
                    "percentC": 0
                }
            }
        
        # Ordenar por valor total decrescente
        items.sort(key=lambda x: x["valor_total"], reverse=True)
        
        # Calcular total e percentuais acumulados
        total_value = sum(item["valor_total"] for item in items)
        accumulated = 0
        
        for item in items:
            accumulated += item["valor_total"]
            accumulated_percentage = (accumulated / total_value * 100) if total_value > 0 else 0
            item["accumulated_percentage"] = round(accumulated_percentage, 1)
            
            # Classificar segundo Pareto (80-15-5)
            if accumulated_percentage <= 80:
                item["classification"] = "A"
            elif accumulated_percentage <= 95:
                item["classification"] = "B"
            else:
                item["classification"] = "C"
        
        # Calcular resumo
        countA = sum(1 for item in items if item["classification"] == "A")
        countB = sum(1 for item in items if item["classification"] == "B")
        countC = sum(1 for item in items if item["classification"] == "C")
        
        valueA = sum(item["valor_total"] for item in items if item["classification"] == "A")
        valueB = sum(item["valor_total"] for item in items if item["classification"] == "B")
        valueC = sum(item["valor_total"] for item in items if item["classification"] == "C")
        
        summary = {
            "total": total_value,
            "countA": countA,
            "countB": countB,
            "countC": countC,
            "valueA": valueA,
            "valueB": valueB,
            "valueC": valueC,
            "percentA": round((valueA / total_value * 100), 1) if total_value > 0 else 0,
            "percentB": round((valueB / total_value * 100), 1) if total_value > 0 else 0,
            "percentC": round((valueC / total_value * 100), 1) if total_value > 0 else 0,
        }
        
        return {
            "status": "success",
            "items": items,
            "summary": summary
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao calcular Curva ABC: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao calcular Curva ABC: {str(e)}",
        )

# ============== EXPORT XLSX ==============
@app.post("/api/export-xlsx")
async def export_xlsx(items: List[Dict]):
    """
    Exporta itens da planilha para arquivo XLSX
    
    Args:
        items: Lista de itens [
            {
                "id": 1,
                "code": "001",
                "description": "Item",
                "unit": "un",
                "qty": 10,
                "unitPrice": 100.00
            }
        ]
    
    Returns:
        arquivo XLSX para download
    """
    try:
        if not Workbook:
            raise HTTPException(
                status_code=500,
                detail="openpyxl não está instalado",
            )
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Orçamento"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        total_font = Font(bold=True, size=11)
        currency_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalho
        headers = ["Código", "Descrição", "Unidade", "Quantidade", "Valor Unitário", "Total"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Dados
        total_geral = 0
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1).value = item.get("code", "")
            ws.cell(row=row_num, column=2).value = item.get("description", "")
            ws.cell(row=row_num, column=3).value = item.get("unit", "")
            
            qty = float(item.get("qty", 0))
            unit_price = float(item.get("unitPrice", 0))
            total = qty * unit_price
            total_geral += total
            
            ws.cell(row=row_num, column=4).value = qty
            ws.cell(row=row_num, column=5).value = unit_price
            ws.cell(row=row_num, column=6).value = total
            
            # Formato moeda para R$
            ws.cell(row=row_num, column=5).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=6).number_format = 'R$ #,##0.00'
            ws.cell(row=row_num, column=6).fill = currency_fill
            
            # Bordas
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = border
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="right" if col >= 4 else "left")
        
        # Linha de Total
        total_row = len(items) + 3
        ws.cell(row=total_row, column=5).value = "TOTAL GERAL:"
        ws.cell(row=total_row, column=5).font = total_font
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
        
        ws.cell(row=total_row, column=6).value = total_geral
        ws.cell(row=total_row, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=total_row, column=6).fill = total_fill
        ws.cell(row=total_row, column=6).font = total_font
        ws.cell(row=total_row, column=6).border = border
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Altura do cabeçalho
        ws.row_dimensions[1].height = 25
        
        # Salvar arquivo
        filename = f"orcamento_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = TEMP_FOLDER / filename
        wb.save(file_path)
        
        logger.info(f"✅ XLSX gerado: {file_path}")
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao exportar XLSX: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao gerar Excel: {str(e)}",
        )

# ============== SERVE FRONTEND INDEX (SPA FALLBACK) ==============
@app.get("/{path_name:path}")
async def serve_frontend(path_name: str):
    """
    Serve o frontend para rotas que não são API
    Necessário para SPA (Single Page Application)
    """
    # Se é um arquivo com extensão (CSS, JS, etc), tentar servir como estático
    if "." in path_name and not path_name.startswith("api"):
        return {"error": "File not found"}
    
    # Servir index.html para rotas do frontend
    index_path = FRONTEND_DIST / "index.html"
    if index_path.exists():
        return FileResponse(index_path)
    
    return {"error": "Frontend not available"}

# ============== RUN ==============
if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=port,
        reload=False,
    )
