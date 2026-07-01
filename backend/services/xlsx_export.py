"""
Geração de XLSX com abas condicionais conforme modelosSelecionados.
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Tuple

from services.analitico_normalize import (
    build_sintetico_grupo_rows,
    classify_tipo_linha,
    normalize_hierarchical_analitico,
    parse_numeric,
)
from services.openai_service import _coerce_bdi, _coerce_number

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore

DEFAULT_MODELS = {
    "analitico": False,
    "sintetico": False,
    "curva_abc": True,
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ANALITICO_GROUP_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GROUP_FONT = Font(bold=True, size=10)
COMP_FONT = Font(italic=True, size=9, color="475569")
TOTAL_FILL = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
ZEBRA_LIGHT = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
ZEBRA_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CLASS_FILLS = {
    "A": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "B": PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid"),
    "C": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
}
CLASS_FONTS = {
    "A": Font(bold=True, color="991B1B"),
    "B": Font(bold=True, color="854D0E"),
    "C": Font(bold=True, color="065F46"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def normalize_models_selection(raw: Dict[str, bool] | None) -> Dict[str, bool]:
    if not raw:
        return dict(DEFAULT_MODELS)
    return {
        "analitico": bool(raw.get("analitico", raw.get("analítico", True))),
        "sintetico": bool(raw.get("sintetico", raw.get("sintético", True))),
        "curva_abc": bool(raw.get("curva_abc", raw.get("abc", True))),
    }


def _bdi_factor(bdi_percent: float) -> float:
    return 1.0 + (bdi_percent / 100.0) if bdi_percent > 0 else 1.0


def _resolve_tipo_linha(raw: Dict[str, Any]) -> str:
    tipo = str(raw.get("tipo_linha") or raw.get("tipo") or "item").strip().lower()
    if tipo in ("grupo", "titulo", "título", "title"):
        return "grupo"
    if tipo in ("composicao", "composição", "insumo", "subitem"):
        return "composicao"
    return "item"


def _is_group_row(raw: Dict[str, Any]) -> bool:
    tipo = _resolve_tipo_linha(raw)
    desc = str(raw.get("description") or raw.get("descricao") or "").strip().lower()
    return tipo == "grupo" or "total do grupo" in desc


def _is_analitico_grupo_visual(row_data: Dict[str, Any]) -> bool:
    """Grupo no Excel: sem Código e sem Unidade (cinza + negrito)."""
    if _is_group_row(row_data):
        return True
    codigo = str(row_data.get("code") or row_data.get("codigo") or "").strip()
    unidade = str(row_data.get("unit") or row_data.get("unidade") or "").strip()
    return not codigo and not unidade


def _is_composicao_row(raw: Dict[str, Any]) -> bool:
    return _resolve_tipo_linha(raw) == "composicao"


def _is_executive_row(raw: Dict[str, Any]) -> bool:
    return _resolve_tipo_linha(raw) == "item" and not _is_group_row(raw)


def prepare_hierarchical_analitico_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Preserva ordem, tipagem, totais e numeração hierárquica (Python, não IA)."""
    payload: List[Dict[str, Any]] = []
    for idx, raw in enumerate(items):
        if not isinstance(raw, dict):
            continue
        payload.append(
            {
                **raw,
                "item_numero": str(
                    raw.get("item_numero") or raw.get("item") or raw.get("id") or ""
                ).strip(),
                "descricao": str(raw.get("descricao") or raw.get("description") or "").strip(),
                "codigo": str(raw.get("codigo") or raw.get("code") or "").strip(),
                "unidade": str(raw.get("unidade") or raw.get("unit") or "").strip(),
                "quantidade": _coerce_number(raw.get("quantidade") or raw.get("qty")),
                "valor_unitario": _coerce_number(
                    raw.get("valor_unitario")
                    or raw.get("unitPrice")
                    or raw.get("unit_com_bdi")
                ),
                "valor_total": _coerce_number(
                    raw.get("valor_total") or raw.get("total_com_bdi") or raw.get("totalValue")
                ),
                "bdi": _coerce_bdi(raw.get("bdi") or raw.get("BDI")),
                "tipo_linha": raw.get("tipo_linha") or raw.get("tipo"),
                "rotulo_linha": str(raw.get("rotulo_linha") or "").strip(),
                "tipo_categoria": str(raw.get("tipo_categoria") or "").strip(),
                "porcentagem": _coerce_number(raw.get("porcentagem") or raw.get("percentual")),
                "banco": str(raw.get("banco") or "").strip(),
                "_order": idx,
            }
        )

    normalized = normalize_hierarchical_analitico(payload)
    rows: List[Dict[str, Any]] = []
    for row in normalized:
        base = _normalize_base_row(row)
        base["item_numero"] = str(row.get("item_numero") or row.get("item") or "").strip()
        base["rotulo_linha"] = str(row.get("rotulo_linha") or "").strip()
        base["tipo_categoria"] = str(row.get("tipo_categoria") or "").strip()
        base["porcentagem"] = _coerce_number(row.get("porcentagem") or 0)
        tipo_linha = str(row.get("tipo_linha") or row.get("tipo") or "").strip().lower()
        if not tipo_linha or tipo_linha in ("none", "null"):
            tipo_linha = classify_tipo_linha(row)
        base["tipo_linha"] = tipo_linha
        base["tipo"] = tipo_linha
        base["banco"] = str(row.get("banco") or "").strip()
        base["qty"] = _coerce_number(row.get("quantidade") or row.get("qty"))
        base["unit"] = str(row.get("unidade") or row.get("unit") or "").strip()
        base["unit_com_bdi"] = _coerce_number(
            row.get("valor_unitario") or row.get("unit_com_bdi")
        )
        total_linha = parse_numeric(row.get("valor_total") or row.get("total_com_bdi"))
        if total_linha <= 0:
            total_linha = parse_numeric(base.get("total_com_bdi"))
        base["total_com_bdi"] = total_linha
        base["valor_total"] = total_linha
        base["totalValue"] = total_linha
        base["code"] = str(row.get("codigo") or row.get("code") or "").strip()
        base["description"] = str(row.get("descricao") or row.get("description") or "").strip()
        base["_order"] = row.get("_order")
        rows.append(base)
    return rows


def _line_total_com_bdi(raw: Dict[str, Any]) -> float:
    bdi = _coerce_bdi(raw.get("bdi") or raw.get("BDI"))
    qty = _coerce_number(raw.get("qty") or raw.get("quantidade") or raw.get("quantity"))
    unit_com_bdi = _coerce_number(
        raw.get("unitPrice") or raw.get("valor_unitario") or raw.get("unitValue")
    )
    total_com_bdi = _coerce_number(
        raw.get("lineTotal")
        or raw.get("line_total")
        or raw.get("totalValue")
        or raw.get("valor_total")
    )
    if total_com_bdi <= 0 and qty > 0 and unit_com_bdi > 0:
        total_com_bdi = qty * unit_com_bdi
    if total_com_bdi <= 0 and qty > 0:
        unit_sem = unit_com_bdi / _bdi_factor(bdi) if _bdi_factor(bdi) > 0 else unit_com_bdi
        total_com_bdi = qty * unit_com_bdi if unit_com_bdi > 0 else qty * unit_sem * _bdi_factor(bdi)
    return total_com_bdi


def _normalize_base_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    bdi = _coerce_bdi(raw.get("bdi") or raw.get("BDI"))
    qty = _coerce_number(raw.get("qty") or raw.get("quantidade") or raw.get("quantity"))
    unit_com_bdi = _coerce_number(
        raw.get("unitPrice") or raw.get("valor_unitario") or raw.get("unitValue")
    )
    total_com_bdi = _line_total_com_bdi(raw)
    factor = _bdi_factor(bdi)
    unit_sem_bdi = unit_com_bdi / factor if factor > 0 else unit_com_bdi
    if unit_com_bdi <= 0 and qty > 0 and total_com_bdi > 0:
        unit_com_bdi = total_com_bdi / qty
    if unit_sem_bdi <= 0 and qty > 0 and total_com_bdi > 0:
        unit_sem_bdi = (total_com_bdi / factor) / qty if factor > 0 else total_com_bdi / qty

    return {
        "code": str(raw.get("code") or raw.get("codigo") or "").strip(),
        "description": str(raw.get("description") or raw.get("descricao") or "").strip(),
        "bdi": bdi,
        "unit": str(raw.get("unit") or raw.get("unidade") or "").strip(),
        "qty": qty,
        "unit_com_bdi": unit_com_bdi,
        "total_com_bdi": total_com_bdi,
        "grupo": str(raw.get("grupo") or "").strip(),
        "classification": str(raw.get("classification") or raw.get("class") or "")
        .strip()
        .upper(),
        "accumulated_percentage": raw.get("accumulated_percentage"),
        "_order": raw.get("_order"),
    }


def prepare_analitico_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Ordem original (campo item/id); apenas itens executivos; sem ABC."""
    rows: List[Dict[str, Any]] = []
    for idx, raw in enumerate(items):
        if not isinstance(raw, dict) or not _is_executive_row(raw):
            continue
        row = _normalize_base_row(raw)
        row["_order"] = _coerce_number(raw.get("item") or raw.get("id") or idx)
        rows.append(row)
    rows.sort(key=lambda r: (r["_order"], r["code"]))
    return rows


def _extract_catalog_code(raw: Dict[str, Any]) -> str:
    catalog = str(
        raw.get("catalogCode")
        or raw.get("catalog_code")
        or raw.get("codigo_catalogo")
        or ""
    ).strip()
    item_numero = str(raw.get("item") or raw.get("item_numero") or "").strip()
    codigo = str(raw.get("codigo") or "").strip()
    if catalog:
        return catalog
    if codigo and codigo != item_numero:
        return codigo
    return ""


def prepare_curva_abc_rows(items: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], float]:
    """Ordenação por valor decrescente + percentuais e classificação A/B/C."""
    prepared: List[Dict[str, Any]] = []
    for raw in items:
        if not isinstance(raw, dict) or not _is_executive_row(raw):
            continue
        row = _normalize_base_row(raw)
        row["item_numero"] = str(raw.get("item") or raw.get("item_numero") or "").strip()
        row["banco"] = str(raw.get("banco") or "").strip()
        row["catalog_code"] = _extract_catalog_code(raw)
        prepared.append(row)

    prepared.sort(key=lambda row: row["total_com_bdi"], reverse=True)
    total_geral = sum(row["total_com_bdi"] for row in prepared)

    accumulated = 0.0
    for row in prepared:
        percent = (row["total_com_bdi"] / total_geral * 100.0) if total_geral > 0 else 0.0
        pct_before = accumulated
        accumulated += percent
        row["percent"] = percent
        acc_front = row.get("accumulated_percentage")
        row["accumulated"] = (
            float(acc_front) if acc_front is not None and acc_front != "" else accumulated
        )
        if not row["classification"]:
            if pct_before < 80:
                row["classification"] = "A"
            elif pct_before < 95:
                row["classification"] = "B"
            else:
                row["classification"] = "C"

    return prepared, total_geral


def prepare_sintetico_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Apenas grupos pai legítimos com totais consolidados (float estrito)."""
    return build_sintetico_grupo_rows(items)


def _write_header_row(ws, headers: List[str]) -> None:
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _apply_col_widths(ws, widths: Dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_novacap_metadata_header(
    ws,
    *,
    nome_obra: str | None = None,
    bancos_referencia: str | None = None,
    bdi_percent: float | None = None,
    encargos_sociais: str | None = None,
) -> int:
    """Cabeçalho estilo NOVACAP (linhas 1-3). Retorna linha inicial dos dados."""
    ws.cell(row=1, column=4).value = "Obra"
    ws.cell(row=1, column=5).value = "Bancos"
    ws.cell(row=1, column=7).value = "B.D.I."
    ws.cell(row=1, column=9).value = "Encargos Sociais"
    for col in (4, 5, 7, 9):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=10)

    ws.cell(row=2, column=4).value = nome_obra or "—"
    ws.cell(row=2, column=5).value = bancos_referencia or "SINAPI / SICRO"
    ws.cell(row=2, column=7).value = (
        f"{bdi_percent:.2f}%".replace(".", ",") if bdi_percent is not None else "—"
    )
    ws.cell(row=2, column=9).value = encargos_sociais or "—"
    for col in (4, 5, 7, 9):
        ws.cell(row=2, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    title_cell = ws.cell(row=3, column=1)
    title_cell.value = "Planilha Orçamentária Analítica"
    title_cell.font = Font(bold=True, size=12)
    return 4


def gerar_aba_analitica(
    ws,
    rows: List[Dict[str, Any]],
    *,
    nome_obra: str | None = None,
    bancos_referencia: str | None = None,
    bdi_percent: float | None = None,
    encargos_sociais: str | None = None,
) -> None:
    """
    Planilha analítica — colunas do edital (A–H):
    Item | Código | Banco | Descrição | Und | Quant. | Valor Unit | Total
    """
    if not bancos_referencia:
        bancos = sorted(
            {
                str(r.get("banco") or "").strip()
                for r in rows
                if str(r.get("banco") or "").strip()
            }
        )
        bancos_referencia = "\n".join(bancos) if bancos else "SINAPI / SICRO"

    if bdi_percent is None:
        bdi_values = [float(r.get("bdi") or 0) for r in rows if float(r.get("bdi") or 0) > 0]
        bdi_percent = sum(bdi_values) / len(bdi_values) if bdi_values else None

    meta_end = _write_novacap_metadata_header(
        ws,
        nome_obra=nome_obra,
        bancos_referencia=bancos_referencia,
        bdi_percent=bdi_percent,
        encargos_sociais=encargos_sociais,
    )

    header_row = meta_end
    headers = [
        "Item",
        "Código",
        "Tipo",
        "Banco",
        "Descrição",
        "Und",
        "Quant.",
        "Valor Unit",
        "Total",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    data_start = header_row + 1
    open_grupos: List[Tuple[int, int]] = []

    def _close_grupo(grupo_row: int, child_start: int, child_end: int) -> None:
        if child_end < child_start:
            return
        total_col = get_column_letter(9)
        total_cell = ws.cell(row=grupo_row, column=9)
        total_cell.value = f"=SUM({total_col}{child_start}:{total_col}{child_end})"
        total_cell.number_format = "#,##0.00"
        total_cell.font = GROUP_FONT
        total_cell.alignment = Alignment(horizontal="right", vertical="center")

    for idx, row_data in enumerate(rows):
        row_num = data_start + idx
        tipo = str(row_data.get("tipo_linha") or "item").lower()
        is_grupo = _is_analitico_grupo_visual(row_data)
        is_comp = tipo == "composicao" and not is_grupo

        item_num = str(row_data.get("item_numero") or "").strip()
        rotulo = str(row_data.get("rotulo_linha") or "").strip()
        col_a = item_num or rotulo
        codigo = str(row_data.get("code") or row_data.get("codigo") or "").strip()
        tipo_label = "Grupo" if is_grupo else ("Composição" if is_comp else "Item")
        banco = str(row_data.get("banco") or "").strip()
        descricao = str(row_data.get("description") or row_data.get("descricao") or "").strip()
        unidade = str(row_data.get("unit") or row_data.get("unidade") or "").strip()
        qty = _coerce_number(row_data.get("qty") or row_data.get("quantidade"))
        bdi = _coerce_bdi(row_data.get("bdi") or row_data.get("BDI"))
        unit_val = _coerce_number(
            row_data.get("unitPrice")
            or row_data.get("valor_unitario")
            or row_data.get("unit_com_bdi")
        )

        row_fill = ANALITICO_GROUP_FILL if is_grupo else (ZEBRA_LIGHT if idx % 2 == 0 else ZEBRA_WHITE)

        if is_grupo:
            if open_grupos:
                prev_row, child_start = open_grupos.pop()
                _close_grupo(prev_row, child_start, row_num - 1)
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = ANALITICO_GROUP_FILL
                cell.font = GROUP_FONT
                cell.border = THIN_BORDER
            ws.cell(row=row_num, column=1).value = col_a
            ws.cell(row=row_num, column=5).value = descricao
            total_cell = ws.cell(row=row_num, column=9)
            total_cell.font = GROUP_FONT
            total_cell.fill = ANALITICO_GROUP_FILL
            total_cell.border = THIN_BORDER
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            open_grupos.append((row_num, row_num + 1))
            continue

        values = {
            1: col_a,
            2: codigo,
            3: tipo_label,
            4: banco,
            5: descricao,
            6: unidade,
            7: qty if qty else None,
            8: unit_val if unit_val else None,
        }

        for col_num, value in values.items():
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = THIN_BORDER
            cell.fill = row_fill
            cell.value = value
            if col_num == 7:
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 8:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 5 and is_comp:
                cell.font = COMP_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            elif col_num == 5:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_num == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        total_cell = ws.cell(row=row_num, column=9)
        total_cell.border = THIN_BORDER
        total_cell.fill = row_fill
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.number_format = "#,##0.00"
        qty_col = get_column_letter(7)
        unit_col = get_column_letter(8)
        static_total = _coerce_number(
            row_data.get("total_com_bdi") or row_data.get("valor_total") or row_data.get("totalValue")
        )
        if qty > 0 and unit_val > 0:
            computed_total = round(qty * unit_val, 2)
            total_cell.value = computed_total if computed_total > 0 else static_total
        elif static_total > 0:
            total_cell.value = static_total

    last_row = data_start + len(rows) - 1
    while open_grupos:
        grupo_row, child_start = open_grupos.pop()
        _close_grupo(grupo_row, child_start, last_row)

    ws.freeze_panes = f"A{data_start}"
    _apply_col_widths(
        ws,
        {"A": 14, "B": 14, "C": 12, "D": 12, "E": 48, "F": 8, "G": 12, "H": 14, "I": 16},
    )


def _fill_analitico_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    gerar_aba_analitica(ws, rows)


def _fill_curva_abc_sheet(
    ws,
    rows: List[Dict[str, Any]],
    total_geral: float,
    *,
    nome_projeto: str | None = None,
) -> None:
    from datetime import datetime

    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    subtitle_font = Font(size=10, color="64748B", italic=True)

    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Curva ABC — Análise de Pareto"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    subtitle = nome_projeto or "Orçamento de Obras"
    ws["A2"].value = f"{subtitle} · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    header_row = 4
    headers = [
        "Item",
        "Banco",
        "Código",
        "Descrição",
        "BDI (%)",
        "Un.",
        "Quantidade",
        "V. Unit. C/BDI",
        "V. Total C/BDI",
        "% Individual",
        "% Acumulado",
        "Classe",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = f"A{header_row + 1}"

    right_cols = {5, 7, 8, 9, 10, 11}
    center_cols = {1, 2, 3, 6, 12}

    for idx, row_data in enumerate(rows):
        row_num = header_row + 1 + idx
        stripe = ZEBRA_LIGHT if idx % 2 == 0 else ZEBRA_WHITE
        cls = str(row_data.get("classification") or "").strip().upper()
        row_fill = CLASS_FILLS.get(cls, stripe)

        values = [
            row_data.get("item_numero") or row_data.get("code") or "",
            row_data.get("banco") or "",
            row_data.get("catalog_code") or "",
            row_data.get("description") or "",
            row_data.get("bdi", 0),
            row_data.get("unit") or "",
            row_data.get("qty", 0),
            row_data.get("unit_com_bdi", 0),
            row_data.get("total_com_bdi", 0),
            row_data.get("percent", 0),
            row_data.get("accumulated", 0),
            cls,
        ]
        formats = [
            None,
            None,
            None,
            None,
            '0.00"%"',
            None,
            "#,##0.0000",
            "#,##0.000",
            "#,##0.00",
            '0.00"%"',
            '0.00"%"',
            None,
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = THIN_BORDER
            if formats[col_num - 1]:
                cell.number_format = formats[col_num - 1]
            if col_num == 12:
                cell.fill = CLASS_FILLS.get(cls, row_fill)
                cell.font = CLASS_FONTS.get(cls, Font(bold=True))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 4:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.fill = row_fill
                if col_num in right_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_num in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    data_end = header_row + len(rows)
    total_row = data_end + 2
    ws.cell(row=total_row, column=8).value = "TOTAL GERAL:"
    ws.cell(row=total_row, column=8).font = TOTAL_FONT
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=9)
    total_cell.value = total_geral
    total_cell.number_format = "#,##0.00"
    total_cell.fill = TOTAL_FILL
    total_cell.font = TOTAL_FONT
    total_cell.border = THIN_BORDER

    summary_start = total_row + 2
    ws.cell(row=summary_start, column=1).value = "Resumo por classe"
    ws.cell(row=summary_start, column=1).font = Font(bold=True, size=11, color="1F4E78")

    summary_headers = ["Classe", "Qtd. Itens", "Valor Total", "% do Total"]
    for col_offset, label in enumerate(summary_headers):
        cell = ws.cell(row=summary_start + 1, column=1 + col_offset)
        cell.value = label
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    class_stats: Dict[str, Dict[str, float]] = {
        "A": {"count": 0, "total": 0.0},
        "B": {"count": 0, "total": 0.0},
        "C": {"count": 0, "total": 0.0},
    }
    for row_data in rows:
        cls = str(row_data.get("classification") or "C").strip().upper()
        if cls not in class_stats:
            cls = "C"
        class_stats[cls]["count"] += 1
        class_stats[cls]["total"] += float(row_data.get("total_com_bdi") or 0)

    for offset, cls in enumerate(("A", "B", "C")):
        row_num = summary_start + 2 + offset
        stats = class_stats[cls]
        pct = (stats["total"] / total_geral * 100.0) if total_geral > 0 else 0.0
        values = [cls, int(stats["count"]), stats["total"], pct]
        formats = [None, "0", "#,##0.00", '0.00"%"']
        for col_offset, (value, fmt) in enumerate(zip(values, formats)):
            cell = ws.cell(row=row_num, column=1 + col_offset)
            cell.value = value
            cell.border = THIN_BORDER
            cell.fill = CLASS_FILLS.get(cls, ZEBRA_WHITE)
            cell.font = CLASS_FONTS.get(cls, Font())
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(
                horizontal="right" if col_offset > 0 else "center",
                vertical="center",
            )

    _apply_col_widths(
        ws,
        {
            "A": 10,
            "B": 16,
            "C": 14,
            "D": 46,
            "E": 9,
            "F": 8,
            "G": 12,
            "H": 14,
            "I": 16,
            "J": 12,
            "K": 12,
            "L": 9,
        },
    )


def gerar_aba_sintetica(
    ws,
    rows: List[Dict[str, Any]],
    *,
    nome_obra: str | None = None,
) -> None:
    """Orçamento Sintético — Item | Descrição do Grupo | Valor Total C/ BDI."""
    meta_end = _write_novacap_metadata_header(ws, nome_obra=nome_obra)
    title_cell = ws.cell(row=3, column=1)
    title_cell.value = "Orçamento Sintético — Resumo Gerencial"
    title_cell.font = Font(bold=True, size=12)

    header_row = meta_end
    headers = ["Item", "Descrição do Grupo", "Valor Total C/ BDI"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    data_start = header_row + 1
    total_geral = 0.0

    for idx, row_data in enumerate(rows):
        row_num = data_start + idx
        item_num = str(row_data.get("item_numero") or row_data.get("item") or "").strip()
        descricao = str(
            row_data.get("descricao") or row_data.get("description") or ""
        ).strip()
        total_val = parse_numeric(
            row_data.get("valor_total") or row_data.get("total_com_bdi") or 0
        )
        total_geral += total_val

        values = [item_num, descricao, total_val]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = GROUP_FONT
            cell.fill = ANALITICO_GROUP_FILL
            cell.border = THIN_BORDER
            if col_num == 3:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    total_row = data_start + len(rows) + 1
    ws.cell(row=total_row, column=2).value = "TOTAL GERAL:"
    ws.cell(row=total_row, column=2).font = TOTAL_FONT
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=3)
    total_cell.value = total_geral
    total_cell.number_format = "#,##0.00"
    total_cell.fill = TOTAL_FILL
    total_cell.font = TOTAL_FONT
    total_cell.border = THIN_BORDER

    ws.freeze_panes = f"A{data_start}"
    _apply_col_widths(ws, {"A": 14, "B": 56, "C": 20})


def _fill_sintetico_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    gerar_aba_sintetica(ws, rows)


def build_export_workbook(
    items: List[Dict[str, Any]],
    modelos_selecionados: Dict[str, bool] | None,
    *,
    nome_projeto: str | None = None,
) -> Tuple[Any, List[str]]:
    if not Workbook:
        raise RuntimeError("openpyxl não está instalado")

    models = normalize_models_selection(modelos_selecionados)
    if not any(models.values()):
        models = dict(DEFAULT_MODELS)

    wb = Workbook()
    wb.remove(wb.active)
    sheets_created: List[str] = []

    if models.get("analitico"):
        analitico_rows = prepare_hierarchical_analitico_rows(items)
        if analitico_rows:
            ws = wb.create_sheet("Orçamento Analítico")
            gerar_aba_analitica(ws, analitico_rows, nome_obra=nome_projeto)
            sheets_created.append("Orçamento Analítico")

    if models.get("sintetico"):
        sintetico_rows = prepare_sintetico_rows(items)
        if sintetico_rows:
            ws = wb.create_sheet("Orçamento Sintético")
            gerar_aba_sintetica(ws, sintetico_rows, nome_obra=nome_projeto)
            sheets_created.append("Orçamento Sintético")

    if models.get("curva_abc"):
        abc_rows, total_geral = prepare_curva_abc_rows(items)
        if abc_rows:
            ws = wb.create_sheet("Curva ABC")
            _fill_curva_abc_sheet(ws, abc_rows, total_geral, nome_projeto=nome_projeto)
            sheets_created.append("Curva ABC")

    if not sheets_created:
        raise ValueError(
            "Nenhuma aba pôde ser gerada. Verifique os itens e os modelos selecionados."
        )

    return wb, sheets_created


# ---------------------------------------------------------------------------
# Extended export: templates SINAPI / Livre, comparativo e formatação padrão
# ---------------------------------------------------------------------------

SINAPI_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
SINAPI_HEADER_FONT = Font(bold=True, color="FFFFFF", size=14)
SINAPI_ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SINAPI_TOTAL_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
COMPARE_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COMPARE_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

LIVRE_COLUMN_LABELS: Dict[str, str] = {
    "codigo": "Código",
    "code": "Código",
    "codigo_sinapi": "Código SINAPI",
    "descricao": "Descrição",
    "description": "Descrição",
    "unidade": "Unidade",
    "unit": "Unidade",
    "quantidade": "Quantidade",
    "qty": "Quantidade",
    "quantity": "Quantidade",
    "precoUnitario": "Preço Unitário",
    "preco_unitario": "Preço Unitário",
    "valor_unitario": "Preço Unitário",
    "unitPrice": "Preço Unitário",
    "unitValue": "Preço Unitário",
    "precoTotal": "Preço Total",
    "preco_total": "Preço Total",
    "valor_total": "Preço Total",
    "totalValue": "Preço Total",
    "lineTotal": "Preço Total",
    "bdi": "BDI (%)",
    "classification": "Classe ABC",
    "class": "Classe ABC",
    "grupo": "Grupo",
    "banco": "Banco",
    "tipo": "Tipo",
}

DEFAULT_LIVRE_COLUMNS = [
    "descricao",
    "unidade",
    "quantidade",
    "precoUnitario",
    "precoTotal",
]

SINAPI_HEADERS = [
    "Código SINAPI",
    "Descrição",
    "Unidade",
    "Quantidade",
    "Preço Unitário",
    "Preço Total",
    "BDI (%)",
    "Total com BDI",
]


def _field_value_for_livre(row: Dict[str, Any], field: str) -> Any:
    aliases: Dict[str, List[str]] = {
        "descricao": ["descricao", "description"],
        "codigo": ["codigo", "code"],
        "codigo_sinapi": ["codigo_sinapi", "codigo", "code", "banco"],
        "unidade": ["unidade", "unit"],
        "quantidade": ["quantidade", "qty", "quantity"],
        "precoUnitario": ["precoUnitario", "preco_unitario", "valor_unitario", "unitPrice", "unitValue"],
        "precoTotal": ["precoTotal", "preco_total", "valor_total", "totalValue", "lineTotal", "total_com_bdi"],
        "bdi": ["bdi", "BDI"],
        "classification": ["classification", "class"],
        "grupo": ["grupo"],
        "banco": ["banco"],
        "tipo": ["tipo", "tipo_linha"],
    }
    keys = aliases.get(field, [field])
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return ""


def _auto_width_columns(ws, max_width: int = 50) -> None:
    if not get_column_letter:
        return
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _apply_sheet_standards(
    ws,
    *,
    header_row: int = 1,
    value_cols: List[int] | None = None,
    qty_cols: List[int] | None = None,
    pct_cols: List[int] | None = None,
) -> None:
    if not get_column_letter:
        return
    last_col = ws.max_column or 1
    last_row = ws.max_row or header_row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    if last_row > header_row:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(last_col)}{last_row}"
        )
    for col in value_cols or []:
        for row in range(header_row + 1, last_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #.##0,00'
    for col in qty_cols or []:
        for row in range(header_row + 1, last_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#.##0,000"
    for col in pct_cols or []:
        for row in range(header_row + 1, last_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0,00%"
    _auto_width_columns(ws)


def _executive_rows_from_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows, _ = prepare_curva_abc_rows(items)
    if rows:
        return rows
    prepared: List[Dict[str, Any]] = []
    for raw in items:
        if isinstance(raw, dict) and _is_executive_row(raw):
            prepared.append(_normalize_base_row(raw))
    return prepared


def gerar_aba_sinapi(ws, items: List[Dict[str, Any]]) -> None:
    ws.title = "Planilha SINAPI"
    rows = _executive_rows_from_items(items)

    for col_num, header in enumerate(SINAPI_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = SINAPI_HEADER_FONT
        cell.fill = SINAPI_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    total_com_bdi = 0.0
    for idx, row_data in enumerate(rows):
        row_num = idx + 2
        stripe = SINAPI_ALT_FILL if idx % 2 else ZEBRA_WHITE
        bdi = _coerce_bdi(row_data.get("bdi"))
        qty = _coerce_number(row_data.get("qty"))
        unit_sem = _coerce_number(row_data.get("unit_com_bdi"))
        if unit_sem <= 0:
            factor = _bdi_factor(bdi)
            unit_sem = _coerce_number(row_data.get("unit_com_bdi")) / factor if factor else 0
        preco_total = qty * unit_sem if qty and unit_sem else 0
        total_linha = preco_total * _bdi_factor(bdi) if preco_total else row_data.get("total_com_bdi", 0)
        total_com_bdi += float(total_linha or 0)

        values = [
            row_data.get("code", ""),
            row_data.get("description", ""),
            row_data.get("unit", ""),
            qty or None,
            unit_sem or None,
            preco_total or None,
            bdi / 100 if bdi else None,
            total_linha or None,
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = stripe
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="right" if col_num >= 4 else "left",
                vertical="center",
                wrap_text=col_num == 2,
            )

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=7).value = "TOTAL"
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=8).value = total_com_bdi
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=8).fill = SINAPI_TOTAL_FILL
    ws.cell(row=total_row, column=8).number_format = 'R$ #.##0,00'

    _apply_sheet_standards(
        ws,
        header_row=1,
        value_cols=[5, 6, 8],
        qty_cols=[4],
        pct_cols=[7],
    )


def gerar_aba_livre(ws, items: List[Dict[str, Any]], colunas: List[str] | None) -> None:
    ws.title = "Exportação Personalizada"
    selected = colunas or DEFAULT_LIVRE_COLUMNS
    if not selected:
        selected = DEFAULT_LIVRE_COLUMNS

    headers = [LIVRE_COLUMN_LABELS.get(c, c) for c in selected]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = _executive_rows_from_items(items)
    if not rows:
        rows = [_normalize_base_row(r) for r in items if isinstance(r, dict)]

    value_cols: List[int] = []
    qty_cols: List[int] = []
    pct_cols: List[int] = []

    for idx, row_data in enumerate(rows):
        row_num = idx + 2
        for col_num, field in enumerate(selected, 1):
            if field in ("precoUnitario", "precoTotal", "preco_unitario", "preco_total", "valor_unitario", "valor_total"):
                value_cols.append(col_num)
            if field in ("quantidade", "qty", "quantity"):
                qty_cols.append(col_num)
            if field in ("bdi",):
                pct_cols.append(col_num)

            raw = dict(row_data)
            if field == "precoUnitario":
                val = _coerce_number(raw.get("unit_com_bdi") or raw.get("valor_unitario"))
            elif field == "precoTotal":
                val = _coerce_number(raw.get("total_com_bdi") or raw.get("valor_total"))
            else:
                val = _field_value_for_livre(raw, field)
            ws.cell(row=row_num, column=col_num).value = val

    _apply_sheet_standards(
        ws,
        header_row=1,
        value_cols=sorted(set(value_cols)),
        qty_cols=sorted(set(qty_cols)),
        pct_cols=sorted(set(pct_cols)),
    )


def _unit_price_from_row(row: Dict[str, Any]) -> float:
    normalized = _normalize_base_row(row)
    unit = _coerce_number(normalized.get("unit_com_bdi"))
    if unit > 0:
        return unit
    qty = _coerce_number(normalized.get("qty"))
    total = _coerce_number(normalized.get("total_com_bdi"))
    return total / qty if qty > 0 else 0.0


def _description_key(row: Dict[str, Any]) -> str:
    return str(row.get("description") or row.get("descricao") or "").strip().lower()


def gerar_abas_comparativo(
    wb,
    budgets: List[Dict[str, Any]],
) -> None:
    """budgets: [{upload_id, nome, items}, ...] até 3."""
    if not budgets or len(budgets) < 2:
        return

    budget_rows: List[Tuple[str, List[Dict[str, Any]]]] = []
    for i, budget in enumerate(budgets[:3]):
        items = budget.get("items") or []
        rows = _executive_rows_from_items(items)
        sheet_name = f"Orçamento {i + 1}"[:31]
        ws = wb.create_sheet(sheet_name)
        headers = ["Código", "Descrição", "Unidade", "Qtd.", "Preço Unit.", "Total", "Classe"]
        _write_header_row(ws, headers)
        for idx, row_data in enumerate(rows):
            row_num = idx + 2
            values = [
                row_data.get("code", ""),
                row_data.get("description", ""),
                row_data.get("unit", ""),
                row_data.get("qty", 0),
                row_data.get("unit_com_bdi", 0),
                row_data.get("total_com_bdi", 0),
                row_data.get("classification", ""),
            ]
            for col_num, value in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num).value = value
        _apply_sheet_standards(ws, header_row=1, value_cols=[5, 6], qty_cols=[4])
        budget_rows.append((sheet_name, rows))

    ws_cmp = wb.create_sheet("Comparativo")
    cmp_headers = [
        "Descrição",
        "Orç. 1 — P. Unit.",
        "Orç. 2 — P. Unit.",
        "Orç. 3 — P. Unit.",
        "Dif. % vs Orç. 1",
        "Menor preço",
    ]
    for col_num, header in enumerate(cmp_headers, 1):
        cell = ws_cmp.cell(row=1, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    price_maps: List[Dict[str, float]] = []
    all_descs: set[str] = set()
    for _, rows in budget_rows:
        pmap: Dict[str, float] = {}
        for row in rows:
            key = _description_key(row)
            if not key:
                continue
            all_descs.add(key)
            pmap[key] = _unit_price_from_row(row)
        price_maps.append(pmap)

    sorted_descs = sorted(all_descs)
    for idx, desc_key in enumerate(sorted_descs):
        row_num = idx + 2
        display_desc = desc_key
        for _, rows in budget_rows:
            for row in rows:
                if _description_key(row) == desc_key:
                    display_desc = str(row.get("description") or row.get("descricao") or desc_key)
                    break

        prices: List[float | None] = []
        for pmap in price_maps:
            val = pmap.get(desc_key)
            prices.append(val if val and val > 0 else None)

        ws_cmp.cell(row=row_num, column=1).value = display_desc
        min_price: float | None = None
        valid_prices = [p for p in prices if p is not None and p > 0]
        if valid_prices:
            min_price = min(valid_prices)

        for col_offset, price in enumerate(prices):
            cell = ws_cmp.cell(row=row_num, column=2 + col_offset)
            if price is None:
                cell.value = "-"
            else:
                cell.value = price
                cell.number_format = 'R$ #.##0,00'
                if min_price is not None and price > min_price * 1.2:
                    cell.fill = COMPARE_RED_FILL

        p1 = prices[0]
        others = [p for p in prices[1:] if p is not None and p > 0]
        diff_cell = ws_cmp.cell(row=row_num, column=5)
        if p1 and others:
            avg_other = sum(others) / len(others)
            diff_pct = ((p1 - avg_other) / avg_other) if avg_other else 0
            diff_cell.value = diff_pct
            diff_cell.number_format = "0,00%"
        else:
            diff_cell.value = "-"

        min_cell = ws_cmp.cell(row=row_num, column=6)
        if min_price is not None:
            min_cell.value = min_price
            min_cell.number_format = 'R$ #.##0,00'
            min_cell.fill = COMPARE_GREEN_FILL
        else:
            min_cell.value = "-"

    _apply_sheet_standards(ws_cmp, header_row=1, value_cols=[2, 3, 4, 6], pct_cols=[5])


def build_export_workbook_extended(
    items: List[Dict[str, Any]],
    modelos_selecionados: Dict[str, bool] | None,
    *,
    nome_projeto: str | None = None,
    template: str = "novacap",
    colunas: List[str] | None = None,
    compare_budgets: List[Dict[str, Any]] | None = None,
) -> Tuple[Any, List[str]]:
    template_key = (template or "novacap").strip().lower()
    sheets_created: List[str] = []

    if template_key == "sinapi":
        if not Workbook:
            raise RuntimeError("openpyxl não está instalado")
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha SINAPI"
        gerar_aba_sinapi(ws, items)
        sheets_created.append("Planilha SINAPI")
    elif template_key == "livre":
        if not Workbook:
            raise RuntimeError("openpyxl não está instalado")
        wb = Workbook()
        ws = wb.active
        gerar_aba_livre(ws, items, colunas)
        sheets_created.append("Exportação Personalizada")
    else:
        wb, sheets_created = build_export_workbook(
            items, modelos_selecionados, nome_projeto=nome_projeto
        )

    if compare_budgets and len(compare_budgets) >= 2:
        gerar_abas_comparativo(wb, compare_budgets)
        sheets_created.extend(
            [f"Orçamento {i + 1}" for i in range(min(3, len(compare_budgets)))]
        )
        sheets_created.append("Comparativo")

    if not sheets_created:
        raise ValueError("Nenhuma aba pôde ser gerada.")

    return wb, sheets_created


def save_export_workbook(
    items: List[Dict[str, Any]],
    modelos_selecionados: Dict[str, bool] | None,
    temp_folder: Path,
    nome_projeto: str | None = None,
    template: str = "novacap",
    colunas: List[str] | None = None,
    compare_budgets: List[Dict[str, Any]] | None = None,
) -> Tuple[Path, str]:
    wb, _ = build_export_workbook_extended(
        items,
        modelos_selecionados,
        nome_projeto=nome_projeto,
        template=template,
        colunas=colunas,
        compare_budgets=compare_budgets,
    )
    stem = "orcamento"
    if nome_projeto and nome_projeto.strip():
        safe = re.sub(r"[^\w\s-]", "", nome_projeto.strip(), flags=re.UNICODE)
        safe = re.sub(r"\s+", "_", safe)[:40]
        if safe:
            stem = safe
    filename = f"{stem}_{uuid.uuid4().hex[:8]}.xlsx"
    file_path = temp_folder / filename
    temp_folder.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    return file_path, filename
